
# app.py
# 구매자재팀용 엑셀 변경 히스토리 추적 Streamlit MVP
# 실행: streamlit run app.py

from __future__ import annotations

import io
import re
import sqlite3
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


APP_TITLE = "구매자재팀 Excel History Tracker"
DB_PATH = Path("history_log.db")


# =========================================================
# 1. 기본 설정
# =========================================================

DEFAULT_IMPORTANT_SHEETS = [
    "바이어 PO데이터 ",
    "자재품목리스트",
    "자재 재고 내역",
    "자재입고내역",
    "자재출고내역",
    "생산계획및 실적 데이터 ",
    "구매데이터",
]

# 헤더가 여러 줄이거나 병합셀인 엑셀을 고려해서, 셀 주변 정보를 넓게 탐색함
KEYWORD_RULES = {
    "단가 변경": [
        "unit price", "unit\nprice", "u/p", "u/p2", "price", "단가", "unit  price"
    ],
    "자재명/품목 변경": [
        "material name", "description", "item", "품목", "자재", "material"
    ],
    "공급업체 변경": [
        "supplier", "suppl", "공급", "vendor"
    ],
    "색상 변경": [
        "color", "color-1", "칼라", "색상"
    ],
    "수량/소요량 변경": [
        "qty", "q'ty", "quantity", "cons", "consumption", "ttl cons", "수량", "소요량"
    ],
    "리드타임 변경": [
        "lead time", "leadtime", "lt", "리드타임"
    ],
    "LOSS 변경": [
        "loss", "로스", "불량", "손실"
    ],
    "MOQ 변경": [
        "moq", "minimum order", "최소주문"
    ],
    "MTL/원가 변경": [
        "mtl", "rm", "material cost", "원가", "cost"
    ],
    "FOB/매출 변경": [
        "fob", "amount", "매출", "판매"
    ],
    "납기/일정 변경": [
        "etd", "eta", "delivery", "납기", "일정", "date"
    ],
    "PO/오더 변경": [
        "po", "order", "pr no", "buyer", "오더", "발주"
    ],
}

HIGH_RISK_CATEGORIES = {
    "단가 변경",
    "수량/소요량 변경",
    "LOSS 변경",
    "MOQ 변경",
    "MTL/원가 변경",
    "FOB/매출 변경",
    "리드타임 변경",
}

NUMERIC_IMPACT_CATEGORIES = {
    "단가 변경",
    "수량/소요량 변경",
    "LOSS 변경",
    "MOQ 변경",
    "MTL/원가 변경",
    "FOB/매출 변경",
}


# =========================================================
# 2. DB / 저장 로직
# =========================================================

def init_db() -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            compared_at TEXT,
            user_name TEXT,
            base_file_name TEXT,
            new_file_name TEXT,
            sheet_name TEXT,
            cell TEXT,
            row_index INTEGER,
            col_index INTEGER,
            column_letter TEXT,
            row_label TEXT,
            col_label TEXT,
            category TEXT,
            old_value TEXT,
            new_value TEXT,
            old_formula TEXT,
            new_formula TEXT,
            change_type TEXT,
            risk_level TEXT,
            impact_note TEXT,
            review_status TEXT,
            review_comment TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def append_history(df: pd.DataFrame) -> None:
    if df.empty:
        return
    init_db()
    conn = sqlite3.connect(DB_PATH)
    df.to_sql("history", conn, if_exists="append", index=False)
    conn.close()


def load_history() -> pd.DataFrame:
    init_db()
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql_query("SELECT * FROM history ORDER BY id DESC", conn)
    finally:
        conn.close()
    return df


# =========================================================
# 3. 엑셀 파싱 유틸
# =========================================================

def file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()[:12]


def normalize_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # 엑셀 소수 오차 완화
        return f"{v:.10g}"
    return str(v).strip()


def to_float_or_none(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")


def get_workbook_from_bytes(file_bytes: bytes, data_only: bool = False):
    return load_workbook(io.BytesIO(file_bytes), data_only=data_only, read_only=False)


def get_cell_value(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def get_nearby_text(ws, row: int, col: int, max_header_rows: int = 12) -> str:
    """
    분류 정확도를 높이기 위해 현재 셀의 상단 헤더, 좌측 라벨, 같은 행 텍스트를 함께 탐색.
    """
    texts = []

    # 같은 열의 상단 헤더 후보
    for r in range(1, min(max_header_rows, ws.max_row) + 1):
        v = ws.cell(r, col).value
        if v is not None:
            texts.append(str(v))

    # 같은 행의 좌측 라벨 후보
    for c in range(1, min(col, 12) + 1):
        v = ws.cell(row, c).value
        if v is not None:
            texts.append(str(v))

    # 바로 위 행/왼쪽 셀
    for rr, cc in [(row - 1, col), (row, col - 1), (row - 1, col - 1)]:
        if rr >= 1 and cc >= 1:
            v = ws.cell(rr, cc).value
            if v is not None:
                texts.append(str(v))

    return " | ".join(texts).lower()


def find_row_label(ws, row: int) -> str:
    """
    변경 셀의 행을 설명할 수 있는 왼쪽 핵심값 추정.
    자재명, supplier, color 등이 보통 좌측에 있으므로 앞쪽 8~12개 셀을 묶어서 보여줌.
    """
    vals = []
    for c in range(1, min(ws.max_column, 12) + 1):
        v = ws.cell(row, c).value
        if v not in [None, ""]:
            vals.append(str(v).replace("\n", " ").strip())
    return " / ".join(vals[:5])


def find_col_label(ws, row: int, col: int) -> str:
    vals = []
    for r in range(1, min(row, 12) + 1):
        v = ws.cell(r, col).value
        if v not in [None, ""]:
            vals.append(str(v).replace("\n", " ").strip())
    return " / ".join(vals[-3:])


def classify_change(ws, row: int, col: int, old_v: Any, new_v: Any) -> str:
    if is_formula(old_v) or is_formula(new_v):
        return "수식 변경"

    context = get_nearby_text(ws, row, col)

    for category, keywords in KEYWORD_RULES.items():
        for kw in keywords:
            if kw.lower() in context:
                return category

    old_num = to_float_or_none(old_v)
    new_num = to_float_or_none(new_v)
    if old_num is not None or new_num is not None:
        return "숫자값 변경"

    return "일반값 변경"


def infer_change_type(old_v: Any, new_v: Any) -> str:
    old_empty = normalize_value(old_v) == ""
    new_empty = normalize_value(new_v) == ""

    if old_empty and not new_empty:
        return "신규 입력"
    if not old_empty and new_empty:
        return "삭제"
    if is_formula(old_v) or is_formula(new_v):
        return "수식 변경"
    return "값 변경"


def infer_risk_level(category: str, old_v: Any, new_v: Any) -> str:
    if category == "수식 변경":
        return "상"
    if category in HIGH_RISK_CATEGORIES:
        return "상"

    old_num = to_float_or_none(old_v)
    new_num = to_float_or_none(new_v)
    if old_num is not None and new_num is not None:
        if old_num == 0 and new_num != 0:
            return "중"
        if old_num != 0:
            rate = abs(new_num - old_num) / abs(old_num)
            if rate >= 0.10:
                return "중"
    return "하"


def make_impact_note(category: str, old_v: Any, new_v: Any) -> str:
    old_num = to_float_or_none(old_v)
    new_num = to_float_or_none(new_v)

    if category == "수식 변경":
        return "MTL/FOB/Margin 등 계산 결과가 바뀔 수 있어 확인 필요"

    if old_num is not None and new_num is not None:
        diff = new_num - old_num
        pct = None if old_num == 0 else diff / old_num * 100
        direction = "증가" if diff > 0 else "감소" if diff < 0 else "변화 없음"

        if category in NUMERIC_IMPACT_CATEGORIES:
            if pct is None:
                return f"{category}: {diff:,.4g} {direction}"
            return f"{category}: {diff:,.4g} {direction} ({pct:,.2f}%)"

        if pct is not None:
            return f"수치 변경: {diff:,.4g} {direction} ({pct:,.2f}%)"
        return f"수치 변경: {diff:,.4g} {direction}"

    if category in {"공급업체 변경", "리드타임 변경"}:
        return "조달 안정성/납기 리스크 영향 확인 필요"

    if category in {"자재명/품목 변경", "색상 변경"}:
        return "BOM 또는 자재 매칭 오류 가능성 확인 필요"

    return "변경 사유 확인 필요"


def compare_excel_files(
    base_bytes: bytes,
    new_bytes: bytes,
    base_file_name: str,
    new_file_name: str,
    user_name: str,
    selected_sheets: Optional[List[str]] = None,
    ignore_blank_to_blank: bool = True,
    compare_formulas: bool = True,
    compare_values: bool = True,
) -> pd.DataFrame:
    """
    이전 버전과 현재 버전의 엑셀 파일을 셀 단위로 비교.
    data_only=False: 수식 비교용
    data_only=True: 계산값 비교용. 단, openpyxl은 수식을 재계산하지 않으므로 엑셀에 저장된 캐시값 기준.
    """
    base_wb_formula = get_workbook_from_bytes(base_bytes, data_only=False)
    new_wb_formula = get_workbook_from_bytes(new_bytes, data_only=False)

    base_wb_value = get_workbook_from_bytes(base_bytes, data_only=True)
    new_wb_value = get_workbook_from_bytes(new_bytes, data_only=True)

    common_sheets = [s for s in new_wb_formula.sheetnames if s in base_wb_formula.sheetnames]
    if selected_sheets:
        common_sheets = [s for s in common_sheets if s in selected_sheets]

    records = []
    compared_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for sheet in common_sheets:
        ws_old_f = base_wb_formula[sheet]
        ws_new_f = new_wb_formula[sheet]
        ws_old_v = base_wb_value[sheet]
        ws_new_v = new_wb_value[sheet]

        max_row = max(ws_old_f.max_row, ws_new_f.max_row)
        max_col = max(ws_old_f.max_column, ws_new_f.max_column)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                old_formula = get_cell_value(ws_old_f, row, col)
                new_formula = get_cell_value(ws_new_f, row, col)
                old_value = get_cell_value(ws_old_v, row, col)
                new_value = get_cell_value(ws_new_v, row, col)

                changed = False

                # 수식 자체 변경 비교
                if compare_formulas and normalize_value(old_formula) != normalize_value(new_formula):
                    changed = True

                # 값 변경 비교
                # 수식셀은 formula와 value가 혼재될 수 있어, formula 비교를 우선함
                if compare_values and not changed:
                    if normalize_value(old_value) != normalize_value(new_value):
                        changed = True

                if not changed:
                    continue

                if ignore_blank_to_blank:
                    if normalize_value(old_formula) == "" and normalize_value(new_formula) == "":
                        if normalize_value(old_value) == "" and normalize_value(new_value) == "":
                            continue

                category = classify_change(ws_new_f, row, col, old_formula, new_formula)
                change_type = infer_change_type(old_formula, new_formula)
                risk = infer_risk_level(category, old_formula, new_formula)
                impact = make_impact_note(category, old_formula, new_formula)

                records.append(
                    {
                        "compared_at": compared_at,
                        "user_name": user_name,
                        "base_file_name": base_file_name,
                        "new_file_name": new_file_name,
                        "sheet_name": sheet,
                        "cell": f"{get_column_letter(col)}{row}",
                        "row_index": row,
                        "col_index": col,
                        "column_letter": get_column_letter(col),
                        "row_label": find_row_label(ws_new_f, row),
                        "col_label": find_col_label(ws_new_f, row, col),
                        "category": category,
                        "old_value": normalize_value(old_value),
                        "new_value": normalize_value(new_value),
                        "old_formula": normalize_value(old_formula) if is_formula(old_formula) else "",
                        "new_formula": normalize_value(new_formula) if is_formula(new_formula) else "",
                        "change_type": change_type,
                        "risk_level": risk,
                        "impact_note": impact,
                        "review_status": "미확인",
                        "review_comment": "",
                    }
                )

    return pd.DataFrame(records)


# =========================================================
# 4. Google Sheets 확장용 어댑터 자리
# =========================================================

class DataSourceAdapter:
    """
    지금은 Excel 업로드 방식.
    나중에 Google Sheets API를 붙일 때 이 클래스만 확장하면 됨.

    예:
    - ExcelUploadAdapter: 업로드 파일 bytes 반환
    - GoogleSheetsAdapter: spreadsheet_id, worksheet_name으로 데이터/수식 반환
    """
    def load(self):
        raise NotImplementedError


class ExcelUploadAdapter(DataSourceAdapter):
    def __init__(self, uploaded_file):
        self.uploaded_file = uploaded_file

    def load(self) -> Tuple[bytes, str]:
        return self.uploaded_file.getvalue(), self.uploaded_file.name


# =========================================================
# 5. UI
# =========================================================

def render_header() -> None:
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="📊",
        layout="wide",
    )

    st.markdown(
        """
        <style>
        .main .block-container {
            padding-top: 1.5rem;
            padding-bottom: 2rem;
        }
        .metric-card {
            padding: 1rem;
            border-radius: 14px;
            border: 1px solid #e5e7eb;
            background: #ffffff;
        }
        .small-caption {
            color: #6b7280;
            font-size: 0.88rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title("📊 구매자재팀 Excel History Tracker")
    st.caption("이전 버전과 현재 버전의 자재 관련 엑셀을 비교하여 변경 이력, 영향 항목, 확인 상태를 관리하는 MVP입니다.")


def render_sidebar() -> Dict[str, Any]:
    st.sidebar.header("설정")

    user_name = st.sidebar.text_input("수정자/검토자 이름", value="구매자재팀")

    st.sidebar.subheader("비교 옵션")
    compare_formulas = st.sidebar.checkbox("수식 변경 비교", value=True)
    compare_values = st.sidebar.checkbox("값 변경 비교", value=True)
    ignore_blank_to_blank = st.sidebar.checkbox("빈 셀 변화 무시", value=True)

    st.sidebar.subheader("시트 선택")
    sheet_mode = st.sidebar.radio(
        "비교 범위",
        ["전체 시트", "주요 시트만", "직접 선택"],
        index=1,
    )

    selected_sheets = None
    if sheet_mode == "주요 시트만":
        selected_sheets = DEFAULT_IMPORTANT_SHEETS
    elif sheet_mode == "직접 선택":
        selected_sheets = st.sidebar.multiselect(
            "비교할 시트명",
            DEFAULT_IMPORTANT_SHEETS,
            default=DEFAULT_IMPORTANT_SHEETS,
        )

    st.sidebar.subheader("저장 옵션")
    save_to_db = st.sidebar.checkbox("비교 결과를 로컬 DB에 저장", value=True)

    return {
        "user_name": user_name,
        "compare_formulas": compare_formulas,
        "compare_values": compare_values,
        "ignore_blank_to_blank": ignore_blank_to_blank,
        "selected_sheets": selected_sheets,
        "save_to_db": save_to_db,
    }


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="history")
    return output.getvalue()


def render_summary(df: pd.DataFrame) -> None:
    if df.empty:
        st.info("변경 사항이 없습니다.")
        return

    total = len(df)
    high = int((df["risk_level"] == "상").sum())
    formula = int((df["category"] == "수식 변경").sum())
    sheets = df["sheet_name"].nunique()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("총 변경 셀", f"{total:,}")
    c2.metric("상 리스크", f"{high:,}")
    c3.metric("수식 변경", f"{formula:,}")
    c4.metric("변경 발생 시트", f"{sheets:,}")

    st.subheader("변경 유형 요약")
    c5, c6 = st.columns([1, 1])

    with c5:
        category_count = df["category"].value_counts().reset_index()
        category_count.columns = ["category", "count"]
        st.dataframe(category_count, use_container_width=True, hide_index=True)

    with c6:
        sheet_count = df["sheet_name"].value_counts().reset_index()
        sheet_count.columns = ["sheet_name", "count"]
        st.dataframe(sheet_count, use_container_width=True, hide_index=True)


def render_filters(df: pd.DataFrame) -> pd.DataFrame:
    st.subheader("변경 이력 상세 조회")

    f1, f2, f3, f4 = st.columns(4)

    with f1:
        sheet_filter = st.multiselect(
            "시트",
            sorted(df["sheet_name"].dropna().unique().tolist()),
            default=[],
        )

    with f2:
        category_filter = st.multiselect(
            "변경 구분",
            sorted(df["category"].dropna().unique().tolist()),
            default=[],
        )

    with f3:
        risk_filter = st.multiselect(
            "리스크",
            ["상", "중", "하"],
            default=[],
        )

    with f4:
        keyword = st.text_input("검색어", placeholder="자재명, 셀, 값, 영향 메모 등")

    filtered = df.copy()

    if sheet_filter:
        filtered = filtered[filtered["sheet_name"].isin(sheet_filter)]
    if category_filter:
        filtered = filtered[filtered["category"].isin(category_filter)]
    if risk_filter:
        filtered = filtered[filtered["risk_level"].isin(risk_filter)]

    if keyword:
        key = keyword.lower()
        search_cols = [
            "sheet_name", "cell", "row_label", "col_label", "category",
            "old_value", "new_value", "old_formula", "new_formula", "impact_note"
        ]
        mask = pd.Series(False, index=filtered.index)
        for col in search_cols:
            if col in filtered.columns:
                mask = mask | filtered[col].astype(str).str.lower().str.contains(key, na=False)
        filtered = filtered[mask]

    return filtered


def render_review_editor(df: pd.DataFrame) -> pd.DataFrame:
    display_cols = [
        "risk_level", "category", "sheet_name", "cell",
        "row_label", "col_label",
        "old_value", "new_value",
        "old_formula", "new_formula",
        "impact_note", "review_status", "review_comment",
    ]

    existing_cols = [c for c in display_cols if c in df.columns]

    edited = st.data_editor(
        df[existing_cols],
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "review_status": st.column_config.SelectboxColumn(
                "review_status",
                help="구매자재팀 확인 상태",
                options=["미확인", "확인 완료", "재검토 필요", "반영 완료"],
                required=True,
            ),
            "review_comment": st.column_config.TextColumn(
                "review_comment",
                help="변경 사유 또는 검토 의견",
            ),
        },
    )
    return edited


def render_history_db_tab() -> None:
    st.subheader("누적 히스토리 DB")
    hist = load_history()

    if hist.empty:
        st.info("아직 저장된 히스토리가 없습니다.")
        return

    st.caption(f"총 {len(hist):,}건 저장됨")

    f1, f2, f3 = st.columns(3)
    with f1:
        hist_sheet = st.multiselect("DB 시트 필터", sorted(hist["sheet_name"].dropna().unique().tolist()))
    with f2:
        hist_cat = st.multiselect("DB 변경 구분 필터", sorted(hist["category"].dropna().unique().tolist()))
    with f3:
        hist_risk = st.multiselect("DB 리스크 필터", ["상", "중", "하"])

    filtered = hist.copy()
    if hist_sheet:
        filtered = filtered[filtered["sheet_name"].isin(hist_sheet)]
    if hist_cat:
        filtered = filtered[filtered["category"].isin(hist_cat)]
    if hist_risk:
        filtered = filtered[filtered["risk_level"].isin(hist_risk)]

    st.dataframe(filtered, use_container_width=True, hide_index=True)

    st.download_button(
        "누적 히스토리 Excel 다운로드",
        data=df_to_excel_bytes(filtered),
        file_name=f"history_db_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def main() -> None:
    render_header()
    settings = render_sidebar()

    tab_compare, tab_history, tab_guide = st.tabs(["파일 비교", "누적 히스토리", "사용 가이드"])

    with tab_compare:
        st.subheader("1. 엑셀 파일 업로드")

        c1, c2 = st.columns(2)
        with c1:
            base_file = st.file_uploader(
                "이전 버전 엑셀",
                type=["xlsx", "xlsm"],
                key="base_file",
            )
        with c2:
            new_file = st.file_uploader(
                "현재 버전 엑셀",
                type=["xlsx", "xlsm"],
                key="new_file",
            )

        st.caption("현재 MVP는 두 엑셀 파일을 비교합니다. 향후 Google Sheets API를 연결하면 실시간 스프레드시트 변경 이력 방식으로 확장할 수 있습니다.")

        if base_file and new_file:
            base_bytes, base_name = ExcelUploadAdapter(base_file).load()
            new_bytes, new_name = ExcelUploadAdapter(new_file).load()

            with st.expander("파일 정보", expanded=False):
                st.write(
                    {
                        "base_file": base_name,
                        "base_hash": file_hash(base_bytes),
                        "new_file": new_name,
                        "new_hash": file_hash(new_bytes),
                    }
                )

            if st.button("변경 이력 비교 실행", type="primary"):
                with st.spinner("엑셀 파일을 비교하는 중입니다..."):
                    try:
                        result_df = compare_excel_files(
                            base_bytes=base_bytes,
                            new_bytes=new_bytes,
                            base_file_name=base_name,
                            new_file_name=new_name,
                            user_name=settings["user_name"],
                            selected_sheets=settings["selected_sheets"],
                            ignore_blank_to_blank=settings["ignore_blank_to_blank"],
                            compare_formulas=settings["compare_formulas"],
                            compare_values=settings["compare_values"],
                        )

                        st.session_state["last_result_df"] = result_df

                        if settings["save_to_db"] and not result_df.empty:
                            append_history(result_df)

                        st.success("비교가 완료되었습니다.")

                    except Exception as e:
                        st.error(f"비교 중 오류가 발생했습니다: {e}")

        if "last_result_df" in st.session_state:
            result_df = st.session_state["last_result_df"]

            st.divider()
            render_summary(result_df)

            if not result_df.empty:
                filtered = render_filters(result_df)

                st.caption(f"필터 적용 결과: {len(filtered):,}건")
                edited = render_review_editor(filtered)

                st.download_button(
                    "현재 필터 결과 Excel 다운로드",
                    data=df_to_excel_bytes(filtered),
                    file_name=f"excel_change_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                csv_bytes = filtered.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
                st.download_button(
                    "현재 필터 결과 CSV 다운로드",
                    data=csv_bytes,
                    file_name=f"excel_change_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                )

    with tab_history:
        render_history_db_tab()

    with tab_guide:
        st.subheader("사용 가이드")
        st.markdown(
            """
            ### 목적
            구매자재팀이 BOM, 자재 단가, 소요량, LOSS, MOQ, MTL 계산식 등 자재 관련 엑셀 변경 사항을 추적하기 위한 MVP입니다.

            ### 현재 기능
            1. 이전 버전 / 현재 버전 엑셀 파일 업로드
            2. 시트별 셀 변경 자동 탐지
            3. 변경 항목 자동 분류
            4. 상/중/하 리스크 분류
            5. MTL, FOB, 단가, 소요량 등 영향 메모 자동 생성
            6. 변경 이력 Excel/CSV 다운로드
            7. 로컬 SQLite DB에 누적 저장

            ### 향후 확장 방향
            - Google Sheets API 연결
            - 수정자 이메일 기반 기록
            - 변경 즉시 Slack/메일 알림
            - MTL/FOB/Margin 영향 자동 재계산
            - ERP/MRP 데이터 연동

            ### 권장 운영 방식
            - 매주 또는 주요 PO 변경 시점마다 최신 파일을 업로드
            - 직전 버전과 비교
            - 상 리스크 항목부터 구매자재팀이 확인
            - review_comment에 변경 사유 기록
            """
        )


if __name__ == "__main__":
    main()
